#!/usr/bin/env python
import sys
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table

def removeSpaces(line):
    return "_".join(line.split(" "))

def padZero(i):
    return ("00" + str(i))[-3:]

if len(sys.argv) != 3:
    print("Usage: python start.py [STUDY_NAME] [EMAIL]")
    sys.exit(1)

# Define prefixes
STUDY_NAME = removeSpaces(sys.argv[1])
USER_EMAIL = sys.argv[2]

std_uri = "seva-kb:STD-{}".format(STUDY_NAME)

dpl_pre = "seva-kb:DPL-CEA-{}-".format(STUDY_NAME)
plt_pre = "seva-kb:PLT-CEA-{}-".format(STUDY_NAME)
ins_pre = "seva-kb:INS-{}-Instrument-".format(STUDY_NAME)
det_model_pre = "seva:DETYP-".format(STUDY_NAME)
det_pre = "seva-kb:DET-{}-".format(STUDY_NAME)

# Create oas directory if necessary
if not os.path.exists('oas'):
    os.makedirs('oas')

# Load instruments and platforms
fp = open("instruments.csv")

instruments = []

for i, l in enumerate(fp):
    if i == 0: continue

    spl = l.rstrip()
    spl = spl.replace("\n", "")
    spl = spl.split(",")

    instruments.append({"instrument": spl[0], "platform": spl[1]})

fp.close()

# Load detectors
detectors = []

fp = open("detectors.csv")
for i, l in enumerate(fp):
    if(i == 0): continue

    spl = l.rstrip()
    spl = spl.replace("\n", "")
    spl = spl.split(",")

    detectors.append({"detector": spl[0], "instrument": spl[1]})

fp.close()

# Load identifiers
identifiers = []

fp = open("identifiers.csv")
for i, l in enumerate(fp):
    if(i == 0): continue

    spl = l.rstrip()
    spl = spl.replace("\n", "")
    spl = spl.split(",")

    identifiers.append({"detector": spl[0], "identifier": spl[1], "unit": spl[2]})

fp.close()

# Load DPL template
wb = load_workbook('templates/DPL-template.xlsx')

# Write deployments
ws = wb["Deployments"]

for i, ins in enumerate(instruments):

    data = [dpl_pre + removeSpaces(ins["instrument"])]
    data.append("vstoi:Deployment")
    data.append(plt_pre + removeSpaces(ins["platform"]))
    data.append(ins_pre + padZero(i + 1))

    filtered_detectors = list(filter(lambda x: x["instrument"] == ins["instrument"], detectors))
    data.append(",".join(map(lambda x: det_pre + x["detector"] + "-" + padZero(i + 1), filtered_detectors)))
    data.append("1900-01-01T00:00:00.000Z")

    ws.append(data)

# Write platforms
ws = wb["Platforms"]

platforms = list(set(map(lambda x: x["platform"], instruments)))

for plt in platforms:

    data = [plt_pre + removeSpaces(plt)]
    data.append("seva:PLTYP-Location")
    data.append(plt)

    #content = plt_pre + removeSpaces(c) + ","
    #content += "seva:PLTYP-Location,"
    #content += c + ","
    #content += ",,,,,,,,,,"

    ws.append(data)

# Write instruments
ws = wb["Instruments"]

for i, ins in enumerate(instruments):

    data = [ins_pre + padZero(i + 1)]
    data.append("seva:INSTYP-Weather-Station")
    data.append(ins["instrument"])

    ws.append(data)

# Write detector models
ws = wb["DetectorModels"]

detectorModels = list(set(map(lambda x: x["detector"], detectors)))

for detM in detectorModels:

    data = [det_model_pre + detM]
    data.append("vstoi:Detector")
    data.append(detM)

    ws.append(data)

# Write detectors
ws = wb["Detectors"]

insts = list(map(lambda i: i["instrument"], instruments))

for det in detectors:

    i = insts.index(det["instrument"])

    data = [det_pre + det["detector"] + "-" + padZero(i + 1)]
    data.append(det_model_pre + det["detector"])
    data.append(det["detector"])
    data.append("")
    data.append(ins_pre + padZero(i + 1))

    ws.append(data)

wb.save("DPL-{}.xlsx".format(STUDY_NAME))

# Load STD template
wb = load_workbook('templates/STD-template.xlsx')

ws = wb["STD.csv"]

ws["A2"] = STUDY_NAME
ws["B2"] = sys.argv[1]
ws["M2"] = USER_EMAIL

wb.save("STD-{}.xlsx".format(STUDY_NAME))

# Load SSD template
wb = load_workbook('templates/SSD-template.xlsx')

ws = wb["SSD"]

ws["B2"] = std_uri
ws["I3"] = std_uri
ws["I4"] = std_uri
ws["M4"] = len(platforms)

ws = wb["SOC-LOCATIONS"]

for p in platforms:
    ws.append([removeSpaces(p), "sio:Location", "env-1"])

wb.save("SSD-{}.xlsx".format(STUDY_NAME))

# Make OAS files
for ins in instruments:

    clean_ins = removeSpaces(ins["instrument"])

    format_dict = {"study_name": STUDY_NAME, "email": USER_EMAIL, "instrument": clean_ins}

    fp = open("oas/OAS-{study_name}-{instrument}.csv".format(**format_dict), "w")

    fp.write("Study ID,da name,data dict,deployment uri,row scope,cell scope,owner email,permission uri\n")
    fp.write('{study_name},{study_name}-{instrument},SDD-{study_name},seva-kb:DPL-CEA-{study_name}-{instrument},,"<<*, seva-kb:LOC-{instrument}-{study_name}>>",{email},http://seva.be.cea.yale.edu#cea'.format(**format_dict))

    fp.close()

# Write SDD
wb = load_workbook('templates/SDD-template.xlsx')

ws = wb["InfoSheet"]

ws["B2"] = STUDY_NAME
ws["B7"] = STUDY_NAME

ws = wb["Dictionary Mapping"]
for identifier in identifiers:
    data = [identifier["detector"], identifier["identifier"], "??airsample", identifier["unit"], "??measurementtime"]
    ws.append(data)

wb.save("SDD-{}.xlsx".format(STUDY_NAME))

print("Generated study metadata.")
print("Please review your files before ingesting into hadatac.")


